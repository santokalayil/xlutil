import datetime
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.worksheet.table import Table, TableStyleInfo
import numpy as np
import pandas as pd
import re
from typing import Tuple
import string
import json

from ..constants import MAX_COLS, MAX_ROWS


def generate_column_ids():
    # Generate column IDs from A to XFD
    column_ids = []
    for i in range(1, MAX_COLS + 1):  # 16385 = 26 (A-Z) + 26*26 (AA-ZZ) + 26*26*26 (AAA-XFD)
        column_id = ""
        n = i
        while n > 0:
            n, remainder = divmod(n - 1, 26)
            column_id = string.ascii_uppercase[remainder] + column_id
        column_ids.append(column_id)
    return column_ids


def generate_row_ids():
    # Generate row IDs from 1 to 1048576
    row_ids = [i for i in range(1, MAX_ROWS + 1)]
    return row_ids


def extract_cell_n_row_ids(cell_id: str) -> Tuple[str, int]:
    # Define regex pattern with grouping for alphabets and numbers
    pattern = r'([A-Za-z]+)(\d+)'

    # Use regex to extract alphabets and numbers
    match = re.match(pattern, cell_id)

    if match:
        column_id = match.group(1)
        row_id = match.group(2)
        
        # Check if column and row IDs are within the expected range
        if not (column_id.isalpha() and row_id.isdigit() and 1 <= int(row_id) <= MAX_ROWS and len(column_id) <= 3):
            raise ValueError("Invalid cell ID format or out of range.")
        
        return column_id.upper(), int(row_id)
    raise ValueError("Invalid cell ID format.")


def paste(df: pd.DataFrame, table_name: str, at: str, wb:Workbook, sheet_name: str, index=False, overwrite_values=False) -> Workbook:
    if index:
        raise NotImplementedError("Copying dataframe along with `index` is not yet impelemented")
    sheets = [i.title for i in wb.worksheets]
    sheet_idx = sheets.index(sheet_name)
    if sheet_name not in sheets:
        raise ValueError("Sheetname not found in specified workbook")
    ws = wb[sheet_name]
    c, r = extract_cell_n_row_ids(at)
    cs = generate_column_ids()
    rs = generate_row_ids()

    cidx = cs.index(c) 
    cols = df.columns.tolist()
    col_ids = cs[cidx: cidx + len(cols)]

    # adding column names
    col_map = {f"{col_id}{r}": col_name for col_id, col_name in zip(col_ids, cols)}
    for cell_id, cell_value in col_map.items():
        if overwrite_values:
            ws[cell_id] = cell_value
        else:
            if pd.isna(ws[cell_id].value):
                ws[cell_id] = cell_value
            else:
                raise ValueError(f"Value already found in cell '{cell_id}'")

    # row values
    row_id = r
    row_count = 0
    for _, row in df.iterrows():
        row_count += 1
        row_id = r + row_count
        for col, col_id in zip(cols, col_ids):
            paste_loc = f"{col_id}{row_id}"
            if overwrite_values:
                ws[paste_loc] = None if pd.isna(row[col]) else row[col]
            else:
                if pd.isna(ws[paste_loc].value):
                    ws[paste_loc] = None if pd.isna(row[col]) else row[col]
                else:
                    raise ValueError(f"Value already found in cell '{paste_loc}'")
    # setting the above info as table for later recognition
    table_start, table_end = at, paste_loc
    table = Table(displayName=table_name, ref=f"{table_start}:{table_end}")
    # Add a default style with striped rows and banded columns
    style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                        showLastColumn=False, showRowStripes=True, showColumnStripes=True)
    table.tableStyleInfo = style
    ws.add_table(table)  # this method make sure no duplicate table names are added
    wb._sheets[sheet_idx] = ws
    return wb


def worksheet_to_dataframe(ws: Worksheet, headers=False) -> pd.DataFrame:
    data = []
    for row in ws.iter_rows(values_only=True):
        data.append(row)

    # Create DataFrame
    dataframe = pd.DataFrame(data)
    if headers:
        dataframe.columns = dataframe.iloc[0]
        dataframe = dataframe.drop(0)
    return dataframe


def worksheet_range_to_dataframe(ws: Worksheet, start_cell: str, end_cell:str, headers=True) -> pd.DataFrame:
    sc, sr = extract_cell_n_row_ids(start_cell)  # start col and start row
    ec, er = extract_cell_n_row_ids(end_cell)  # end col and end row
    
    column_letters_list = generate_column_ids()
    sc_loc = column_letters_list.index(sc) + 1  # +1 since 1 based index
    ec_loc = column_letters_list.index(ec) + 1  # +1 due to 1 based index
    
    data = []
    for row in ws.iter_rows(min_row=sr, max_row=er, min_col=sc_loc, max_col=ec_loc, values_only=True):
        data.append(row)

    # Create DataFrame
    dataframe = pd.DataFrame(data)
    if headers:
        dataframe.columns = dataframe.iloc[0]
        dataframe = dataframe.drop(0)
    return dataframe

