from pathlib import Path
from typing_extensions import Self
from typing import Any, Dict, Union, Optional
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from .utils import paste, worksheet_to_dataframe

class ExcelSheet:
    "Base Sheet class"


class ExcelFile:
    """Base Excel File class"""
    
    def __init__(self, filepath: Union[Path, str] = None) -> None:
        # self.sheets: Dict = dict()
        if filepath and Path(filepath).is_file():
            self.open(filepath)
            self._filepath = filepath
        else:
            self._workbook: Optional[Workbook] = None
            self._filepath = None
        
    def open(self, filepath: Union[Path, str]) -> None:
        """
        Opens an existing Excel file.

        Args:
            filepath (Union[Path, str]): The path to the Excel file.

        Raises:
            Exception: If the filepath is not a valid Path or str.

        Returns:
            None
        """
        if not isinstance(filepath, (Path, str)):
            raise Exception("filepath should be a pathlib.Path or str")
        self._workbook = load_workbook(filepath, keep_vba=Path(filepath).suffix in [".xlsm", ".XLSM"])
        
    def initialize_new_workbook(self, remove_initial_sheet=True):
        self._workbook = Workbook()
        if remove_initial_sheet is True:
            ws = self._workbook.worksheets[0]
            # Delete the worksheet
            self._workbook.remove(ws)
        
    def new_sheet(self, sheet_name: str) -> Worksheet: 
        if self._workbook is None:
            self.initialize_new_workbook()
        ws = self._workbook.create_sheet(title=sheet_name)
        # this internally adds the worksheet to the workbook
            
        return ws
    
    def add_sheet(
        self, sheet_name: str, dataframe: pd.DataFrame, replace=True, index=True
    ) -> None:
        """Adds new sheet to the excel

        Args:
            sheet_name (str): name of the sheet
            dataframe (pd.DataFrame): data to be saved to the sheet
        """
        if self._workbook is None:
            self.initialize_new_workbook()
        
        if not replace:
            if sheet_name in self.sheets.keys():
                raise Exception(
                    "Unable to add the new file since the sheet_name already exists"
                )

        if not isinstance(index, bool):
            raise ValueError("The index value should be True or False")
        
        if index:
            dataframe = dataframe.reset_index()
        
        if sheet_name in self.sheetnames:
            if not replace:
                raise IndexError("The Sheet name already exists. if you really want to replace, please se the replace=True")
        else:
            self.new_sheet(sheet_name)
           
        self.paste(df=dataframe, table_name=sheet_name, at="A1", sheet_name=sheet_name)
    
    def __setitem__(self, __name: str, __value: pd.DataFrame) -> None:
        if isinstance(__value, pd.DataFrame):
            self.add_sheet(sheet_name=__name, dataframe=__value, replace=True, index=True)
        else:
            raise ValueError("The value to be set should be a pandas dataframe")
    
    def view_sheet(self, sheet_name: str) -> pd.DataFrame:
        """view the sheet as dataframe if sheet name exists

        Args:
            sheet_name (str): name of the sheet

        Raises:
            KeyError: if sheet does not exist, raises error

        Returns:
            pd.DataFrame: sheet as DataFrame
        """
        if sheet_name in self.sheets.keys():
            ws = self._workbook[sheet_name]
            return worksheet_to_dataframe(ws, headers=True)
        else:
            raise KeyError(f"The sheet '{sheet_name}' is not found in the excel object")
    
    def __getitem__(self, __name: str) -> pd.DataFrame:
        return self.view_sheet(__name)

    def __getattr__(self, __name: str) -> pd.DataFrame:
        if __name == "keys":
            return self.__dict__
        return self.view_sheet(__name)
    
    def __delitem__(self, key):
        if key in self.sheets.keys():
            ws = self._workbook[key]
            self._workbook.remove(ws)
            # idx = self._find_index_of_sheet(key)
            # self._workbook._sheets = [s for s in self._workbook._sheets if s.title != key]
        else:
            raise KeyError("The sheet does not exist.")
    
    @property
    def sheetnames(self) -> list:
        return [i.title for i in self._workbook.worksheets]
    
    @property
    def sheets(self) -> dict:
        return {i.title: i for i in self._workbook.worksheets}
    
    def _find_index_of_sheet(self, sheet_name: str) -> int:
        return self.sheetnames.index(sheet_name)
    
    def paste(
            self, 
            df: pd.DataFrame,
            table_name: str,
            at: str,
            sheet_name: str,
            index: bool = False,
            overwrite_values: bool = False
        ) -> None:
        if index is True:
            df = df.reset_index()
        self._workbook = paste(
            df=df,
            table_name=table_name,
            at=at,
            wb=self._workbook,
            sheet_name=sheet_name,
            index=index,
            overwrite_values=overwrite_values,
        )
        
        
    def save(self, filepath: Union[Path, str]= None) -> None:
        """Saves the excel to filesystem

        Args:
            filepath (Path | str): path or filename to save the excel file

        Raises:
            Exception: if filepath is not either path instance or str
        """
        if filepath is None:
            if self._filepath is not None:
                filepath = self._filepath
            else:
                raise Exception("You cannot save file without a filepath")

        if not (isinstance(filepath, Path) or isinstance(filepath, str)):
            raise Exception(
                "The save path should be a pathlib.Path or str formated path"
            )
        
        if isinstance(self._workbook, Workbook):
            self._workbook.save(filepath)
        else:
            raise ValueError("The excel file empty. You cannot save empty file")
        
    def __str__(self) -> str:
        return "Base Excel File"
    
    def __repr__(self) -> pd.DataFrame:
        return repr(str(self))
    
    @property
    def see(self) -> pd.DataFrame:
        sheets: Dict = self.sheets
        if "Sheet1" in sheets.keys():
            return self["Sheet1"].copy()
        elif sheets.keys():
            return sheets[list(sheets.keys())[0]]
        raise ValueError("No sheets are set as value")
    
    
    @staticmethod
    def _load_df_from_path(path: Union[str, Path]) -> pd.DataFrame:
        if isinstance(path, str):
            path = Path(path)
        elif not isinstance(path, Path):
            raise ValueError("Value for path variable is of neither str or Path type")
        if not path.is_file():
            raise ValueError("The path passed is does not exist or not a file")
        if path.suffix.lower() in [".xlsx", ".xls", ".xlsm"]:
            return pd.read_excel(path)
        elif path.suffix.lower() in [".csv"]:
            return pd.read_csv(path)
        elif path.suffix.lower() in [".parquet"]:
            return pd.read_parquet(path)
        elif path.suffix.lower() in [".feather"]:
            return pd.read_feather(path)
        raise TypeError("File type is not supported")
            
            
        
    
    def load_data(self, value: Any) -> None:
        if isinstance(value, pd.DataFrame):
            self.add_sheet("Sheet1", value, replace=False),
        elif isinstance(value, (str, Path)):
            self.add_sheet("Sheet1", self._load_df_from_path(value), replace=False)
        else:
            raise ValueError("The type of value you are trying to set is not supported")
            
            
        


class NewExcelFile:
    """ExcelFile class to create a new excel with multiple sheets."""

    def __init__(self) -> None:
        """Initialize the ExcelFile"""
        self.sheets: Dict[str, pd.DataFrame] = dict()
        self.index_info: Dict[str, bool] = dict()
        self.index = 0

    def add_sheet(
        self, sheet_name: str, dataframe: pd.DataFrame, replace=True, index=True
    ) -> None:
        """Adds new sheet to the excel

        Args:
            sheet_name (str): name of the sheet
            dataframe (pd.DataFrame): data to be saved to the sheet
        """
        if not replace:
            if sheet_name in self.sheets.keys():
                raise Exception(
                    "Unable to add the new file since the sheet_name already exists"
                )

        if not isinstance(index, bool):
            raise ValueError("The index value should be True or False")

        self.index_info[sheet_name] = index
        self.sheets[sheet_name] = dataframe

    def view_sheet(self, sheet_name: str) -> pd.DataFrame:
        """view the sheet as dataframe if sheet name exists

        Args:
            sheet_name (str): name of the sheet

        Raises:
            KeyError: if sheet does not exist, raises error

        Returns:
            pd.DataFrame: sheet as DataFrame
        """
        if sheet_name in self.sheets.keys():
            return self.sheets[sheet_name]
        else:
            raise KeyError(f"The sheet '{sheet_name}' is not found in the excel object")

    def __getitem__(self, __name: str) -> pd.DataFrame:
        return self.view_sheet(__name)

    def __getattr__(self, __name: str) -> pd.DataFrame:
        if __name == "keys":
            return self.__dict__
        return self.view_sheet(__name)

    def __setitem__(self, __name: str, __value: pd.DataFrame) -> None:
        if isinstance(__value, pd.DataFrame):
            self.add_sheet(__name, __value)
        else:
            raise ValueError("The value to be set should be a pandas dataframe")

    def __delitem__(self, key):
        if key in self.sheets.keys():
            del self.sheets[key]
        else:
            raise KeyError("The sheet does not exist.")

    def save(self, filepath: Union[Path, str]) -> None:
        """Saves the excel to filesystem

        Args:
            filepath (Path | str): path or filename to save the excel file

        Raises:
            Exception: if filepath is not either path instance or str
        """

        if not (isinstance(filepath, Path) or isinstance(filepath, str)):
            raise Exception(
                "The save path should be a pathlib.Path or str formated path"
            )

        writer = pd.ExcelWriter(filepath, engine="openpyxl")

        if self.sheets:
            for sheet_name, dataframe in self.sheets.items():
                dataframe.to_excel(
                    writer, sheet_name=sheet_name, index=self.index_info[sheet_name]
                )
        writer.close()

    def __repr__(self) -> str:
        return f"""Excel file with {len(self.sheets)} sheets. 
    Sheets: {", ".join(self.sheets.keys())}"""

    def __iter__(self) -> Self:
        return self

    def __next__(self) -> pd.DataFrame:
        if self.index < len(self.sheets):
            sheet = self.sheets[list(self.sheets.keys())[self.index]]
            self.index += 1
            return sheet
        else:
            self.index = 0
            raise StopIteration

    def __dict__(self) -> Dict[str, pd.DataFrame]:
        return self.sheets.copy()

    def __len__(self) -> int:
        return len(self.sheets)
